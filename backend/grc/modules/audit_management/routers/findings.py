from typing import List, Optional
from datetime import datetime
import json
import logging
import os
import uuid
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, text
from pydantic import BaseModel
from openai import OpenAI

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ....models import (
    AuditFinding, AuditManagementResponse, AuditRecommendation,
    AuditActionPlan, AuditFollowUp, AuditEngagement, Risk,
    GRCUser, get_db
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant

router = APIRouter(prefix="/findings", tags=["Audit - Findings"])
logger = logging.getLogger(__name__)

FINDING_ATTACHMENT_UPLOAD_DIR = os.path.join("backend", "uploads", "audit_findings")
os.makedirs(FINDING_ATTACHMENT_UPLOAD_DIR, exist_ok=True)

ROOT_CAUSE_CATEGORIES = ["people", "process", "technology", "governance"]
SEVERITY_LEVELS = ["critical", "high", "medium", "low", "observation"]


class FindingCreate(BaseModel):
    engagement_id: int
    title: str
    condition: Optional[str] = None
    criteria: Optional[str] = None
    cause: Optional[str] = None
    effect: Optional[str] = None
    root_cause_category: Optional[str] = None
    severity: Optional[str] = "medium"
    framework_mappings: Optional[list] = []
    risk_id: Optional[int] = None
    control_id: Optional[int] = None
    owner_id: Optional[int] = None
    due_date: Optional[datetime] = None
    theme: Optional[str] = None


class FindingUpdate(BaseModel):
    title: Optional[str] = None
    condition: Optional[str] = None
    criteria: Optional[str] = None
    cause: Optional[str] = None
    effect: Optional[str] = None
    root_cause_category: Optional[str] = None
    severity: Optional[str] = None
    status: Optional[str] = None
    framework_mappings: Optional[list] = None
    risk_id: Optional[int] = None
    control_id: Optional[int] = None
    owner_id: Optional[int] = None
    due_date: Optional[datetime] = None
    theme: Optional[str] = None


class ManagementResponseCreate(BaseModel):
    response_type: str
    response_text: Optional[str] = None
    action_plan: Optional[str] = None
    target_date: Optional[datetime] = None


class RecommendationCreate(BaseModel):
    title: str
    description: Optional[str] = None
    priority: Optional[str] = "medium"
    owner_id: Optional[int] = None
    due_date: Optional[datetime] = None


class ActionPlanCreate(BaseModel):
    milestone: str
    description: Optional[str] = None
    owner_id: Optional[int] = None
    due_date: Optional[datetime] = None


class ActionPlanUpdate(BaseModel):
    milestone: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    completed_date: Optional[datetime] = None
    evidence_of_completion: Optional[str] = None


class FollowUpCreate(BaseModel):
    follow_up_type: Optional[str] = "retest"
    retest_result: Optional[str] = None
    retest_details: Optional[str] = None
    evidence_id: Optional[int] = None
    notes: Optional[str] = None


class FollowUpClose(BaseModel):
    notes: Optional[str] = None


class FindingAISuggestRequest(BaseModel):
    engagement_id: Optional[int] = None
    title: str
    condition: Optional[str] = None
    criteria: Optional[str] = None
    cause: Optional[str] = None
    effect: Optional[str] = None
    severity_hint: Optional[str] = None


FINDING_TEMPLATE_HEADERS = [
    "Engagement ID",
    "Title",
    "Condition",
    "Criteria",
    "Cause",
    "Effect",
    "Root Cause Category",
    "Severity",
    "Theme",
    "Due Date",
    "Owner ID",
    "Risk ID",
    "Control ID",
]


def get_openai_client() -> Optional[OpenAI]:
    api_key = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY")
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    if not api_key:
        return None
    kwargs = {"api_key": api_key}
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


def _heuristic_finding_suggestion(title: str, condition: Optional[str], severity_hint: Optional[str]) -> dict:
    text = f"{title} {(condition or '')}".lower()
    severity = (severity_hint or "").lower() or "medium"
    if any(k in text for k in ["breach", "fraud", "unauthorized", "regulatory violation", "critical"]):
        severity = "critical"
    elif any(k in text for k in ["high", "material", "major", "non-compliant"]):
        severity = "high"

    root_cause = "process"
    if any(k in text for k in ["training", "human", "staff", "manual"]):
        root_cause = "people"
    elif any(k in text for k in ["system", "application", "access", "patch", "config", "technology"]):
        root_cause = "technology"
    elif any(k in text for k in ["policy", "oversight", "committee", "approval", "governance"]):
        root_cause = "governance"

    theme = "control deficiency"
    if any(k in text for k in ["access", "identity", "privilege"]):
        theme = "access management"
    elif any(k in text for k in ["vendor", "third party", "supplier"]):
        theme = "third party risk"
    elif any(k in text for k in ["compliance", "regulation", "policy"]):
        theme = "regulatory compliance"

    return {
        "condition": condition or "Observed control breakdown during audit testing.",
        "criteria": "Established policy, procedure, or control standard was not fully met.",
        "cause": "Control design and/or execution gap at process owner level.",
        "effect": "Increased likelihood of operational, compliance, and financial risk.",
        "root_cause_category": root_cause,
        "severity": severity if severity in SEVERITY_LEVELS else "medium",
        "theme": theme,
    }


def suggest_finding_details(data: FindingAISuggestRequest, db: Session, current_user: GRCUser) -> dict:
    base = _heuristic_finding_suggestion(data.title, data.condition, data.severity_hint)
    client = get_openai_client()
    if not client:
        return {**base, "source": "heuristic"}


def _parse_optional_int(value: Optional[object]) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def _parse_optional_date(value: Optional[object]) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except Exception:
        return None


def _normalize_finding_row(row: dict) -> dict:
    def _canon(value: str) -> str:
        return "".join(ch for ch in (value or "").strip().lower() if ch.isalnum())

    alias_map = {
        "engagement id": "engagement_id",
        "engagementid": "engagement_id",
        "finding title": "title",
        "findingtitle": "title",
        "title": "title",
        "finding description": "description",
        "findingdescription": "description",
        "description": "description",
        "severity": "severity",
        "target date": "due_date",
        "targetdate": "due_date",
        "due date": "due_date",
        "duedate": "due_date",
        "audit name": "audit_name",
        "auditname": "audit_name",
        "management response": "management_response",
        "managementresponse": "management_response",
        "status": "status",
        "finding status": "status",
        "findingstatus": "status",
        "related control": "related_control",
        "relatedcontrol": "related_control",
        "finding id": "finding_ref",
        "findingid": "finding_ref",
    }

    normalized = {}
    for key, value in row.items():
        key_text = str(key or "").strip().lower()
        canon = _canon(key_text)
        mapped_key = alias_map.get(key_text) or alias_map.get(canon) or key_text
        normalized[mapped_key] = value

    def get(*keys: str):
        for key in keys:
            if key in normalized and normalized[key] is not None and str(normalized[key]).strip() != "":
                return normalized[key]
        return None

    engagement_raw = get("engagement_id", "engagement id", "engagement")
    engagement_id = _parse_optional_int(engagement_raw)
    title = get("title", "finding title")

    severity_val = str(get("severity") or "medium").strip().lower()
    severity_map = {
        "critical": "critical",
        "high": "high",
        "medium": "medium",
        "low": "low",
        "observation": "observation",
        "info": "observation",
        "informational": "observation",
    }
    severity_val = severity_map.get(severity_val, "medium")

    status_raw = str(get("status") or "").strip().lower().replace(" ", "_")
    status_map = {
        "open": "open",
        "in_progress": "in_progress",
        "inprogress": "in_progress",
        "management_agreed": "management_agreed",
        "managementagreed": "management_agreed",
        "remediated": "remediated",
        "closed": "closed",
    }
    status_val = status_map.get(status_raw) if status_raw else None

    description_text = get("description", "finding description")
    management_response = get("management_response", "management response")
    related_control = get("related_control", "related control")
    criteria_text = ""
    if management_response:
        criteria_text = f"Management Response: {str(management_response).strip()}"
    if related_control:
        criteria_text = (criteria_text + " | " if criteria_text else "") + f"Related Control: {str(related_control).strip()}"

    return {
        "engagement_id": engagement_id,
        "engagement_ref": str(engagement_raw).strip() if engagement_raw is not None else None,
        "audit_name": get("audit_name", "audit name"),
        "title": str(title).strip() if title else "",
        "condition": get("condition") or description_text,
        "criteria": get("criteria") or criteria_text or None,
        "cause": get("cause"),
        "effect": get("effect"),
        "root_cause_category": get("root cause category", "root_cause_category"),
        "severity": severity_val,
        "status": status_val,
        "theme": get("theme"),
        "due_date": _parse_optional_date(get("due date", "due_date")),
        "owner_id": _parse_optional_int(get("owner id", "owner_id")),
        "risk_id": _parse_optional_int(get("risk id", "risk_id")),
        "control_id": _parse_optional_int(get("control id", "control_id")),
    }


def _extract_rows_from_uploaded_finding_file(file_name: str, file_bytes: bytes) -> List[dict]:
    file_name_lower = (file_name or "").lower()
    raw_rows: List[dict] = []

    if file_name_lower.endswith(".csv"):
        import csv
        text = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(text.splitlines())
        for row in reader:
            normalized = {str(k).strip().lower(): v for k, v in (row or {}).items() if k}
            if normalized:
                raw_rows.append(normalized)
    elif file_name_lower.endswith((".xlsx", ".xls")):
        if openpyxl is None:
            raise HTTPException(status_code=500, detail="openpyxl is not installed")
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        best_rows: List[dict] = []
        for ws in wb.worksheets:
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue

            header_idx = None
            for idx, r in enumerate(all_rows[:20]):
                headers = [str(v or "").strip().lower() for v in r]
                if "title" in headers and ("engagement id" in headers or "engagement_id" in headers):
                    header_idx = idx
                    break
            if header_idx is None:
                continue

            headers = [str(v or "").strip().lower() for v in all_rows[header_idx]]
            sheet_rows: List[dict] = []
            for r in all_rows[header_idx + 1:]:
                if not r or not any(v is not None and str(v).strip() for v in r):
                    continue
                item = {}
                for i, value in enumerate(r):
                    if i < len(headers) and headers[i]:
                        item[headers[i]] = value
                if item:
                    sheet_rows.append(item)

            if len(sheet_rows) > len(best_rows):
                best_rows = sheet_rows

        raw_rows = best_rows
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .xlsx/.xls or .csv")

    return [_normalize_finding_row(r) for r in raw_rows]

    user_tenants = get_user_tenants(current_user, db)
    context = ""
    if data.engagement_id and user_tenants:
        engagement = db.query(AuditEngagement).filter(
            AuditEngagement.id == data.engagement_id,
            AuditEngagement.tenant_id.in_(user_tenants)
        ).first()
        if engagement:
            context = f"Engagement title: {engagement.title}\nType: {engagement.engagement_type}\nScope: {engagement.scope or ''}\nObjectives: {engagement.objectives or ''}"

    prompt = f"""You are an internal audit QA assistant.
Given the finding draft below, propose concise audit-quality text.
Return STRICT JSON with keys: condition, criteria, cause, effect, root_cause_category, severity, theme.
Use one of root_cause_category: people, process, technology, governance.
Use one of severity: critical, high, medium, low, observation.

Finding title: {data.title}
Condition: {data.condition or ''}
Criteria: {data.criteria or ''}
Cause: {data.cause or ''}
Effect: {data.effect or ''}
Severity hint: {data.severity_hint or ''}
{context}
"""

    try:
        completion = client.chat.completions.create(
            model="gpt-4o",
            temperature=0.2,
            messages=[
                {"role": "system", "content": "Return valid JSON only."},
                {"role": "user", "content": prompt},
            ],
        )
        content = (completion.choices[0].message.content or "").strip()
        parsed = json.loads(content)
        merged = {
            "condition": parsed.get("condition") or base["condition"],
            "criteria": parsed.get("criteria") or base["criteria"],
            "cause": parsed.get("cause") or base["cause"],
            "effect": parsed.get("effect") or base["effect"],
            "root_cause_category": parsed.get("root_cause_category") or base["root_cause_category"],
            "severity": parsed.get("severity") or base["severity"],
            "theme": parsed.get("theme") or base["theme"],
        }
        if merged["root_cause_category"] not in ROOT_CAUSE_CATEGORIES:
            merged["root_cause_category"] = base["root_cause_category"]
        if merged["severity"] not in SEVERITY_LEVELS:
            merged["severity"] = base["severity"]
        return {**merged, "source": "ai"}
    except Exception as exc:
        logger.warning(f"Finding AI suggestion fallback: {exc}")
        return {**base, "source": "heuristic"}


def ensure_finding_attachment_columns(db: Session):
    bind = db.get_bind()
    dialect = bind.dialect.name if bind is not None else ""

    if dialect == "sqlite":
        cols = db.execute(text("PRAGMA table_info('grc_audit_findings')")).fetchall()
        col_names = {row[1] for row in cols}
        if "attachment_file_name" not in col_names:
            db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN attachment_file_name VARCHAR(255)"))
        if "attachment_file_path" not in col_names:
            db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN attachment_file_path VARCHAR(500)"))
        if "attachment_content_type" not in col_names:
            db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN attachment_content_type VARCHAR(100)"))
        if "attachment_file_size" not in col_names:
            db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN attachment_file_size INTEGER"))
        db.commit()
        return

    db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN IF NOT EXISTS attachment_file_name VARCHAR(255)"))
    db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN IF NOT EXISTS attachment_file_path VARCHAR(500)"))
    db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN IF NOT EXISTS attachment_content_type VARCHAR(100)"))
    db.execute(text("ALTER TABLE grc_audit_findings ADD COLUMN IF NOT EXISTS attachment_file_size INTEGER"))
    db.commit()


def create_finding_record(data: FindingCreate, db: Session, current_user: GRCUser, ai_generated: bool = False) -> AuditFinding:
    user_tenants = get_user_tenants(current_user, db)
    eng = db.query(AuditEngagement).filter(
        AuditEngagement.id == data.engagement_id,
        AuditEngagement.tenant_id.in_(user_tenants)
    ).first()
    if not eng:
        raise HTTPException(status_code=404, detail="Engagement not found")

    count = db.query(AuditFinding).filter(
        AuditFinding.engagement_id == data.engagement_id
    ).count()
    finding_num = f"F-{eng.engagement_number}-{count + 1:03d}" if eng.engagement_number else f"F-{count + 1:03d}"

    finding = AuditFinding(
        tenant_id=eng.tenant_id,
        engagement_id=data.engagement_id,
        finding_number=finding_num,
        title=data.title,
        condition=data.condition,
        criteria=data.criteria,
        cause=data.cause,
        effect=data.effect,
        root_cause_category=data.root_cause_category,
        severity=data.severity,
        framework_mappings=data.framework_mappings or [],
        risk_id=data.risk_id,
        control_id=data.control_id,
        owner_id=data.owner_id,
        due_date=data.due_date,
        theme=data.theme,
        ai_generated=ai_generated,
    )
    db.add(finding)
    db.commit()

    if data.severity in ["critical", "high"] and not data.risk_id:
        risk = Risk(
            tenant_id=eng.tenant_id,
            title=f"Audit Finding: {data.title}",
            description=f"Auto-created from audit finding {finding_num}. Condition: {data.condition or 'N/A'}",
            category="compliance",
            risk_category="audit_finding",
            inherent_likelihood=4 if data.severity == "critical" else 3,
            inherent_impact=4 if data.severity == "critical" else 3,
            inherent_score=16 if data.severity == "critical" else 9,
            residual_likelihood=4 if data.severity == "critical" else 3,
            residual_impact=4 if data.severity == "critical" else 3,
            residual_score=16 if data.severity == "critical" else 9,
            treatment_plan="mitigate",
            status="identified",
            owner_id=data.owner_id,
        )
        db.add(risk)
        db.commit()
        db.refresh(risk)
        finding.risk_id = risk.id
        db.commit()

    db.refresh(finding)
    return finding


def serialize_finding(f: AuditFinding) -> dict:
    responses = []
    if f.management_responses:
        for r in f.management_responses:
            responses.append({
                "id": r.id,
                "response_type": r.response_type,
                "response_text": r.response_text,
                "action_plan": r.action_plan,
                "target_date": r.target_date.isoformat() if r.target_date else None,
                "respondent_id": r.respondent_id,
                "responded_at": r.responded_at.isoformat() if r.responded_at else None,
            })
    
    recs = []
    if f.recommendations:
        for rec in f.recommendations:
            action_plans = []
            if rec.action_plans:
                for ap in rec.action_plans:
                    action_plans.append({
                        "id": ap.id,
                        "milestone": ap.milestone,
                        "description": ap.description,
                        "owner_id": ap.owner_id,
                        "due_date": ap.due_date.isoformat() if ap.due_date else None,
                        "completed_date": ap.completed_date.isoformat() if ap.completed_date else None,
                        "status": ap.status,
                        "evidence_of_completion": ap.evidence_of_completion,
                    })
            recs.append({
                "id": rec.id,
                "title": rec.title,
                "description": rec.description,
                "priority": rec.priority,
                "status": rec.status,
                "owner_id": rec.owner_id,
                "due_date": rec.due_date.isoformat() if rec.due_date else None,
                "action_plans": action_plans,
            })
    
    follow_ups = []
    if f.follow_ups:
        for fu in f.follow_ups:
            follow_ups.append({
                "id": fu.id,
                "follow_up_type": fu.follow_up_type,
                "retest_result": fu.retest_result,
                "retest_details": fu.retest_details,
                "evidence_id": fu.evidence_id,
                "performed_by_id": fu.performed_by_id,
                "performed_at": fu.performed_at.isoformat() if fu.performed_at else None,
                "closure_approved": fu.closure_approved,
                "notes": fu.notes,
            })
    
    return {
        "id": f.id,
        "tenant_id": f.tenant_id,
        "engagement_id": f.engagement_id,
        "engagement_title": f.engagement.title if f.engagement else None,
        "finding_number": f.finding_number,
        "title": f.title,
        "condition": f.condition,
        "criteria": f.criteria,
        "cause": f.cause,
        "effect": f.effect,
        "root_cause_category": f.root_cause_category,
        "severity": f.severity,
        "status": f.status,
        "framework_mappings": f.framework_mappings,
        "risk_id": f.risk_id,
        "control_id": f.control_id,
        "owner_id": f.owner_id,
        "owner_name": (f.owner.display_name or f.owner.username) if f.owner else None,
        "due_date": f.due_date.isoformat() if f.due_date else None,
        "ai_generated": f.ai_generated,
        "theme": f.theme,
        "attachment_file_name": f.attachment_file_name,
        "attachment_file_path": f.attachment_file_path,
        "attachment_content_type": f.attachment_content_type,
        "attachment_file_size": f.attachment_file_size,
        "management_responses": responses,
        "recommendations": recs,
        "follow_ups": follow_ups,
        "created_at": f.created_at.isoformat() if f.created_at else None,
        "updated_at": f.updated_at.isoformat() if f.updated_at else None,
    }


@router.get("")
def list_findings(
    engagement_id: Optional[int] = None,
    severity: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    theme: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"findings": [], "total": 0}
    
    query = db.query(AuditFinding).options(
        joinedload(AuditFinding.management_responses),
        joinedload(AuditFinding.recommendations).joinedload(AuditRecommendation.action_plans),
        joinedload(AuditFinding.follow_ups),
        joinedload(AuditFinding.engagement),
    ).filter(AuditFinding.tenant_id.in_(user_tenants))
    
    if engagement_id:
        query = query.filter(AuditFinding.engagement_id == engagement_id)
    if severity:
        query = query.filter(AuditFinding.severity == severity)
    if status_filter:
        query = query.filter(AuditFinding.status == status_filter)
    if theme:
        query = query.filter(AuditFinding.theme == theme)
    
    findings = query.order_by(AuditFinding.created_at.desc()).all()
    return {"findings": [serialize_finding(f) for f in findings], "total": len(findings)}


@router.get("/grouped-by-engagement")
def get_findings_grouped_by_engagement(
    severity: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"groups": [], "total_engagements": 0, "total_findings": 0}

    query = db.query(AuditFinding).options(
        joinedload(AuditFinding.management_responses),
        joinedload(AuditFinding.recommendations).joinedload(AuditRecommendation.action_plans),
        joinedload(AuditFinding.follow_ups),
        joinedload(AuditFinding.engagement),
    ).filter(AuditFinding.tenant_id.in_(user_tenants))

    if severity:
        query = query.filter(AuditFinding.severity == severity)
    if status_filter:
        query = query.filter(AuditFinding.status == status_filter)

    findings = query.order_by(AuditFinding.created_at.desc()).all()
    grouped = {}
    for finding in findings:
        key = finding.engagement_id
        if key not in grouped:
            grouped[key] = {
                "engagement_id": key,
                "engagement_title": finding.engagement.title if finding.engagement else f"Engagement {key}",
                "total_findings": 0,
                "findings": [],
            }
        grouped[key]["total_findings"] += 1
        grouped[key]["findings"].append(serialize_finding(finding))

    groups = sorted(grouped.values(), key=lambda g: g["total_findings"], reverse=True)
    return {"groups": groups, "total_engagements": len(groups), "total_findings": len(findings)}


@router.get("/template/download")
def download_findings_template(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="No tenant access")

    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FindingsTemplate"
    ws.append(FINDING_TEMPLATE_HEADERS)

    sample_eng = db.query(AuditEngagement).filter(
        AuditEngagement.tenant_id.in_(user_tenants)
    ).order_by(AuditEngagement.created_at.desc()).first()
    sample_engagement_id = sample_eng.id if sample_eng else ""

    ws.append([
        sample_engagement_id,
        "User access review exceptions identified",
        "Quarterly access reviews were not performed for privileged accounts.",
        "Policy requires quarterly privileged access reviews.",
        "Ownership and monitoring controls were not consistently executed.",
        "Increased risk of unauthorized access and policy non-compliance.",
        "process",
        "high",
        "access management",
        datetime.utcnow().strftime("%Y-%m-%d"),
        "",
        "",
        "",
    ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    file_name = f"audit_findings_template_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/import")
async def import_findings_from_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required")

    file_bytes = await file.read()
    rows = _extract_rows_from_uploaded_finding_file(file.filename, file_bytes)
    user_tenants = get_user_tenants(current_user, db)

    created = 0
    skipped = 0
    errors: List[str] = []

    for idx, row in enumerate(rows, start=2):
        try:
            if not row.get("title"):
                skipped += 1
                continue

            engagement = None
            if row.get("engagement_id"):
                engagement = db.query(AuditEngagement).filter(
                    AuditEngagement.id == row["engagement_id"],
                    AuditEngagement.tenant_id.in_(user_tenants)
                ).first()

            if not engagement and row.get("engagement_ref"):
                ref = str(row.get("engagement_ref") or "").strip()
                if ref:
                    engagement = db.query(AuditEngagement).filter(
                        AuditEngagement.engagement_number == ref,
                        AuditEngagement.tenant_id.in_(user_tenants)
                    ).first()

            if not engagement and row.get("audit_name"):
                audit_name = str(row.get("audit_name") or "").strip()
                if audit_name:
                    engagement = db.query(AuditEngagement).filter(
                        AuditEngagement.title.ilike(f"%{audit_name}%"),
                        AuditEngagement.tenant_id.in_(user_tenants)
                    ).first()

            if not engagement:
                skipped += 1
                errors.append(
                    f"Row {idx}: Engagement not found (Engagement ID/Ref: {row.get('engagement_ref') or row.get('engagement_id')}, Audit Name: {row.get('audit_name')})"
                )
                continue

            draft = FindingCreate(
                engagement_id=engagement.id,
                title=row["title"],
                condition=row.get("condition"),
                criteria=row.get("criteria"),
                cause=row.get("cause"),
                effect=row.get("effect"),
                root_cause_category=row.get("root_cause_category"),
                severity=row.get("severity") or "medium",
                risk_id=row.get("risk_id"),
                control_id=row.get("control_id"),
                owner_id=row.get("owner_id"),
                due_date=row.get("due_date"),
                theme=row.get("theme"),
                framework_mappings=[],
            )

            suggestion = suggest_finding_details(
                FindingAISuggestRequest(
                    engagement_id=draft.engagement_id,
                    title=draft.title,
                    condition=draft.condition,
                    criteria=draft.criteria,
                    cause=draft.cause,
                    effect=draft.effect,
                    severity_hint=draft.severity,
                ),
                db,
                current_user,
            ) or {}

            draft = draft.copy(update={
                "condition": draft.condition or suggestion.get("condition"),
                "criteria": draft.criteria or suggestion.get("criteria"),
                "cause": draft.cause or suggestion.get("cause"),
                "effect": draft.effect or suggestion.get("effect"),
                "root_cause_category": draft.root_cause_category or suggestion.get("root_cause_category"),
                "theme": draft.theme or suggestion.get("theme"),
                "severity": draft.severity if draft.severity and draft.severity != "medium" else suggestion.get("severity", draft.severity),
            })

            created_finding = create_finding_record(draft, db, current_user, ai_generated=True)
            if row.get("status") and created_finding:
                created_finding.status = row.get("status")
                created_finding.updated_at = datetime.utcnow()
                db.commit()
            created += 1
        except Exception as exc:
            skipped += 1
            errors.append(f"Row {idx}: {str(exc)}")

    return {
        "message": "Findings import completed",
        "total_rows": len(rows),
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }


@router.post("/ai-suggest")
def ai_suggest_finding_details(
    data: FindingAISuggestRequest,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    return suggest_finding_details(data, db, current_user)


@router.get("/themes")
def get_finding_themes(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"themes": []}

    themes = db.query(
        AuditFinding.theme,
        func.count(AuditFinding.id).label("count"),
    ).filter(
        AuditFinding.tenant_id.in_(user_tenants),
        AuditFinding.theme.isnot(None)
    ).group_by(AuditFinding.theme).all()

    severity_rows = db.query(
        AuditFinding.theme,
        AuditFinding.severity,
    ).filter(
        AuditFinding.tenant_id.in_(user_tenants),
        AuditFinding.theme.isnot(None),
        AuditFinding.severity.isnot(None),
    ).distinct().all()

    severity_map = {}
    for theme, severity in severity_rows:
        if theme not in severity_map:
            severity_map[theme] = []
        if severity not in severity_map[theme]:
            severity_map[theme].append(severity)

    return {"themes": [{
        "theme": t.theme,
        "count": t.count,
        "severities": severity_map.get(t.theme, []),
    } for t in themes]}


@router.get("/overdue")
def get_overdue_findings(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"findings": [], "total": 0}
    
    now = datetime.utcnow()
    findings = db.query(AuditFinding).options(
        joinedload(AuditFinding.engagement),
    ).filter(
        AuditFinding.tenant_id.in_(user_tenants),
        AuditFinding.status.in_(["open", "in_progress", "management_agreed", "management_disagreed", "partially_agreed", "retest_failed"]),
        AuditFinding.due_date < now
    ).order_by(AuditFinding.due_date.asc()).all()
    
    return {"findings": [serialize_finding(f) for f in findings], "total": len(findings)}


@router.post("", status_code=status.HTTP_201_CREATED)
def create_finding(
    data: FindingCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    suggestion = suggest_finding_details(
        FindingAISuggestRequest(
            engagement_id=data.engagement_id,
            title=data.title,
            condition=data.condition,
            criteria=data.criteria,
            cause=data.cause,
            effect=data.effect,
            severity_hint=data.severity,
        ),
        db,
        current_user,
    ) or {}

    filled = data.copy(update={
        "condition": data.condition or suggestion.get("condition"),
        "criteria": data.criteria or suggestion.get("criteria"),
        "cause": data.cause or suggestion.get("cause"),
        "effect": data.effect or suggestion.get("effect"),
        "root_cause_category": data.root_cause_category or suggestion.get("root_cause_category"),
        "theme": data.theme or suggestion.get("theme"),
        "severity": data.severity if data.severity and data.severity != "medium" else suggestion.get("severity", data.severity),
    })

    finding = create_finding_record(filled, db, current_user, ai_generated=True)
    return serialize_finding(finding)


@router.post("/with-attachment", status_code=status.HTTP_201_CREATED)
async def create_finding_with_attachment(
    engagement_id: int = Form(...),
    title: str = Form(...),
    condition: Optional[str] = Form(None),
    criteria: Optional[str] = Form(None),
    cause: Optional[str] = Form(None),
    effect: Optional[str] = Form(None),
    root_cause_category: Optional[str] = Form(None),
    severity: Optional[str] = Form("medium"),
    risk_id: Optional[int] = Form(None),
    control_id: Optional[int] = Form(None),
    owner_id: Optional[int] = Form(None),
    due_date: Optional[str] = Form(None),
    theme: Optional[str] = Form(None),
    attachment: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)

    parsed_due_date: Optional[datetime] = None
    if due_date:
        try:
            parsed_due_date = datetime.fromisoformat(due_date)
        except ValueError:
            try:
                parsed_due_date = datetime.strptime(due_date, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid due_date format. Use YYYY-MM-DD")

    payload = FindingCreate(
        engagement_id=engagement_id,
        title=title,
        condition=condition,
        criteria=criteria,
        cause=cause,
        effect=effect,
        root_cause_category=root_cause_category,
        severity=severity,
        framework_mappings=[],
        risk_id=risk_id,
        control_id=control_id,
        owner_id=owner_id,
        due_date=parsed_due_date,
        theme=theme,
    )

    suggestion = suggest_finding_details(
        FindingAISuggestRequest(
            engagement_id=engagement_id,
            title=title,
            condition=condition,
            criteria=criteria,
            cause=cause,
            effect=effect,
            severity_hint=severity,
        ),
        db,
        current_user,
    ) or {}

    payload = payload.copy(update={
        "condition": payload.condition or suggestion.get("condition"),
        "criteria": payload.criteria or suggestion.get("criteria"),
        "cause": payload.cause or suggestion.get("cause"),
        "effect": payload.effect or suggestion.get("effect"),
        "root_cause_category": payload.root_cause_category or suggestion.get("root_cause_category"),
        "theme": payload.theme or suggestion.get("theme"),
        "severity": payload.severity if payload.severity and payload.severity != "medium" else suggestion.get("severity", payload.severity),
    })

    finding = create_finding_record(payload, db, current_user, ai_generated=True)

    if attachment and attachment.filename:
        file_ext = os.path.splitext(attachment.filename)[1]
        safe_name = f"finding_{finding.id}_{uuid.uuid4().hex}{file_ext}"
        save_path = os.path.join(FINDING_ATTACHMENT_UPLOAD_DIR, safe_name)

        file_bytes = await attachment.read()
        with open(save_path, "wb") as f:
            f.write(file_bytes)

        finding.attachment_file_name = attachment.filename
        finding.attachment_file_path = save_path.replace('\\', '/')
        finding.attachment_content_type = attachment.content_type
        finding.attachment_file_size = len(file_bytes)
        db.commit()
        db.refresh(finding)

    return serialize_finding(finding)


@router.post("/{finding_id}")
async def create_finding_with_attachment_fallback(
    finding_id: str,
    engagement_id: int = Form(...),
    title: str = Form(...),
    condition: Optional[str] = Form(None),
    criteria: Optional[str] = Form(None),
    cause: Optional[str] = Form(None),
    effect: Optional[str] = Form(None),
    root_cause_category: Optional[str] = Form(None),
    severity: Optional[str] = Form("medium"),
    risk_id: Optional[int] = Form(None),
    control_id: Optional[int] = Form(None),
    owner_id: Optional[int] = Form(None),
    due_date: Optional[str] = Form(None),
    theme: Optional[str] = Form(None),
    attachment: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    if finding_id != "with-attachment":
        raise HTTPException(status_code=405, detail="Method not allowed")

    return await create_finding_with_attachment(
        engagement_id=engagement_id,
        title=title,
        condition=condition,
        criteria=criteria,
        cause=cause,
        effect=effect,
        root_cause_category=root_cause_category,
        severity=severity,
        risk_id=risk_id,
        control_id=control_id,
        owner_id=owner_id,
        due_date=due_date,
        theme=theme,
        attachment=attachment,
        db=db,
        current_user=current_user,
    )


@router.get("/{finding_id}")
def get_finding(
    finding_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).options(
        joinedload(AuditFinding.management_responses),
        joinedload(AuditFinding.recommendations).joinedload(AuditRecommendation.action_plans),
        joinedload(AuditFinding.follow_ups),
        joinedload(AuditFinding.engagement),
    ).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    return serialize_finding(finding)


@router.put("/{finding_id}")
def update_finding(
    finding_id: int,
    data: FindingUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    ensure_finding_attachment_columns(db)
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    for field, value in data.dict(exclude_unset=True).items():
        setattr(finding, field, value)
    
    finding.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(finding)
    return serialize_finding(finding)


@router.post("/{finding_id}/management-response", status_code=status.HTTP_201_CREATED)
def add_management_response(
    finding_id: int,
    data: ManagementResponseCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    response = AuditManagementResponse(
        finding_id=finding_id,
        response_type=data.response_type,
        response_text=data.response_text,
        action_plan=data.action_plan,
        target_date=data.target_date,
        respondent_id=current_user.id,
    )
    db.add(response)
    
    if data.response_type == "agree":
        finding.status = "management_agreed"
    elif data.response_type == "disagree":
        finding.status = "management_disagreed"
    elif data.response_type == "partial":
        finding.status = "partially_agreed"
    
    db.commit()
    db.refresh(response)
    return {"id": response.id, "response_type": response.response_type, "message": "Response added"}


@router.post("/{finding_id}/recommendations", status_code=status.HTTP_201_CREATED)
def add_recommendation(
    finding_id: int,
    data: RecommendationCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    rec = AuditRecommendation(
        finding_id=finding_id,
        title=data.title,
        description=data.description,
        priority=data.priority,
        owner_id=data.owner_id,
        due_date=data.due_date,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return {"id": rec.id, "title": rec.title, "message": "Recommendation added"}


@router.post("/{finding_id}/recommendations/{rec_id}/action-plans", status_code=status.HTTP_201_CREATED)
def add_action_plan(
    finding_id: int,
    rec_id: int,
    data: ActionPlanCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    rec = db.query(AuditRecommendation).filter(
        AuditRecommendation.id == rec_id,
        AuditRecommendation.finding_id == finding_id
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Recommendation not found")
    
    ap = AuditActionPlan(
        recommendation_id=rec_id,
        milestone=data.milestone,
        description=data.description,
        owner_id=data.owner_id,
        due_date=data.due_date,
    )
    db.add(ap)
    db.commit()
    db.refresh(ap)
    return {"id": ap.id, "milestone": ap.milestone, "message": "Action plan added"}


@router.put("/{finding_id}/recommendations/{rec_id}/action-plans/{ap_id}")
def update_action_plan(
    finding_id: int,
    rec_id: int,
    ap_id: int,
    data: ActionPlanUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    ap = db.query(AuditActionPlan).filter(
        AuditActionPlan.id == ap_id,
        AuditActionPlan.recommendation_id == rec_id
    ).first()
    if not ap:
        raise HTTPException(status_code=404, detail="Action plan not found")
    
    for field, value in data.dict(exclude_unset=True).items():
        setattr(ap, field, value)
    
    db.commit()
    db.refresh(ap)
    return {"id": ap.id, "status": ap.status, "message": "Action plan updated"}


@router.post("/{finding_id}/follow-ups", status_code=status.HTTP_201_CREATED)
def add_follow_up(
    finding_id: int,
    data: FollowUpCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    fu = AuditFollowUp(
        finding_id=finding_id,
        follow_up_type=data.follow_up_type,
        retest_result=data.retest_result,
        retest_details=data.retest_details,
        evidence_id=data.evidence_id,
        performed_by_id=current_user.id,
        notes=data.notes,
    )
    db.add(fu)
    
    if data.retest_result == "pass":
        finding.status = "remediated"
    elif data.retest_result == "fail":
        finding.status = "retest_failed"
    
    db.commit()
    db.refresh(fu)
    return {"id": fu.id, "retest_result": fu.retest_result, "message": "Follow-up added"}


@router.post("/{finding_id}/follow-ups/{fu_id}/close")
def close_finding(
    finding_id: int,
    fu_id: int,
    data: FollowUpClose,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    fu = db.query(AuditFollowUp).filter(
        AuditFollowUp.id == fu_id,
        AuditFollowUp.finding_id == finding_id
    ).first()
    if not fu:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    
    if fu.retest_result != "pass":
        raise HTTPException(status_code=400, detail="Cannot close finding: follow-up retest result must be 'pass'")

    if finding.status not in ("remediated", "open", "in_progress", "management_agreed", "retest_failed"):
        raise HTTPException(status_code=400, detail=f"Cannot close finding in '{finding.status}' status")

    fu.closure_approved = True
    fu.closure_approved_by_id = current_user.id
    fu.closure_approved_at = datetime.utcnow()
    finding.status = "closed"
    finding.updated_at = datetime.utcnow()
    
    db.commit()
    return {"message": "Finding closed successfully"}


@router.post("/{finding_id}/ai-root-cause")
def ai_suggest_root_cause(
    finding_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).options(
        joinedload(AuditFinding.engagement),
    ).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")

    context_parts = [
        f"Finding: {finding.title}",
        f"Condition: {finding.condition or 'N/A'}",
        f"Criteria: {finding.criteria or 'N/A'}",
        f"Cause: {finding.cause or 'N/A'}",
        f"Effect: {finding.effect or 'N/A'}",
        f"Severity: {finding.severity or 'N/A'}",
        f"Root Cause Category: {finding.root_cause_category or 'N/A'}",
    ]
    if finding.engagement:
        context_parts.append(f"Engagement: {finding.engagement.title}")
        if finding.engagement.scope:
            context_parts.append(f"Scope: {finding.engagement.scope}")

    similar_findings = db.query(AuditFinding).filter(
        AuditFinding.tenant_id.in_(user_tenants),
        AuditFinding.id != finding_id,
        AuditFinding.root_cause_category == finding.root_cause_category,
    ).limit(5).all()

    if similar_findings:
        context_parts.append(f"\nHistorical patterns — {len(similar_findings)} similar findings in same root cause category:")
        for sf in similar_findings:
            context_parts.append(f"  - {sf.title}: {sf.cause or 'No cause documented'}")

    context = "\n".join(context_parts)

    heuristic_causes = []
    text_lower = f"{finding.title} {finding.condition or ''} {finding.cause or ''}".lower()
    if any(k in text_lower for k in ["access", "privilege", "unauthorized"]):
        heuristic_causes.append({"category": "technology", "description": "Inadequate access control mechanisms or privilege management gaps.", "likelihood": "high"})
    if any(k in text_lower for k in ["training", "awareness", "human", "manual"]):
        heuristic_causes.append({"category": "people", "description": "Insufficient staff training, awareness gaps, or over-reliance on manual processes.", "likelihood": "high"})
    if any(k in text_lower for k in ["policy", "procedure", "process", "segregation"]):
        heuristic_causes.append({"category": "process", "description": "Process design deficiency — missing or outdated policies, inadequate segregation of duties.", "likelihood": "medium"})
    if any(k in text_lower for k in ["oversight", "governance", "committee", "monitoring"]):
        heuristic_causes.append({"category": "governance", "description": "Governance and oversight gaps — lack of monitoring or committee review.", "likelihood": "medium"})
    if not heuristic_causes:
        heuristic_causes.append({"category": finding.root_cause_category or "process", "description": "Control design or operational effectiveness gap requiring further analysis.", "likelihood": "medium"})

    client = get_openai_client()
    if not client:
        return {"root_causes": heuristic_causes, "summary": "Heuristic root cause analysis based on finding keywords and patterns.", "source": "heuristic"}

    prompt = f"""You are an internal audit root cause analysis expert.
Analyze the audit finding below and suggest 2-4 potential root causes.
For each root cause, provide: category (people/process/technology/governance), a concise description, and likelihood (high/medium/low).
Also provide a brief summary paragraph of the overall root cause analysis.
Return STRICT JSON with keys: root_causes (array of objects with category, description, likelihood), summary (string).

{context}"""

    try:
        completion = client.chat.completions.create(
            model="gpt-4o",
            temperature=0.3,
            messages=[
                {"role": "system", "content": "Return valid JSON only."},
                {"role": "user", "content": prompt},
            ],
        )
        content = (completion.choices[0].message.content or "").strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[1].rsplit("```", 1)[0].strip()
        parsed = json.loads(content)
        raw_causes = parsed.get("root_causes")
        if not isinstance(raw_causes, list):
            raw_causes = heuristic_causes
        validated_causes = []
        for rc in raw_causes:
            if isinstance(rc, dict) and "description" in rc:
                validated_causes.append({
                    "category": rc.get("category", "process"),
                    "description": str(rc["description"]),
                    "likelihood": rc.get("likelihood", "medium"),
                })
        if not validated_causes:
            validated_causes = heuristic_causes
        return {
            "root_causes": validated_causes,
            "summary": str(parsed.get("summary", "AI-powered root cause analysis.")),
            "source": "ai",
        }
    except Exception as exc:
        logger.warning(f"AI root cause suggestion fallback: {exc}")
        return {"root_causes": heuristic_causes, "summary": "Heuristic root cause analysis (AI unavailable).", "source": "heuristic"}


@router.delete("/{finding_id}")
def delete_finding(
    finding_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    finding = db.query(AuditFinding).filter(
        AuditFinding.id == finding_id,
        AuditFinding.tenant_id.in_(user_tenants)
    ).first()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    
    db.delete(finding)
    db.commit()
    return {"message": "Finding deleted"}
