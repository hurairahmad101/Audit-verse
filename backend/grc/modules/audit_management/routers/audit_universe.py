import os
import csv
import json
import logging
from io import BytesIO
from typing import List, Optional, Dict
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel

try:
    import openpyxl
except ImportError:
    openpyxl = None

from ....models import (
    AuditableEntity, AuditEngagement, AuditFinding, Risk,
    BusinessUnit, GRCUser, get_db
)
from ....routers.auth_router import require_auth, get_user_tenants, get_user_primary_tenant
from ..scoring import (
    apply_score_to_entity, normalize_weights,
    FACTOR_DEFINITIONS, MANUAL_FACTOR_KEYS,
)
from ..lifecycle import rescore_entity
from .scoring import get_or_create_config

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/universe", tags=["Audit - Universe"])


INDUSTRIES = [
    "Banking", "Healthcare", "Insurance", "Technology", "Energy",
    "Government", "Manufacturing", "Retail", "Telecom", "Other",
]


class AuditableEntityCreate(BaseModel):
    name: str
    entity_type: str
    description: Optional[str] = None
    business_unit_id: Optional[int] = None
    audit_cycle_months: Optional[int] = 12
    owner_id: Optional[int] = None
    industry: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_designation: Optional[str] = None
    metadata_json: Optional[dict] = {}
    risk_factors: Optional[Dict[str, float]] = None


class AuditableEntityUpdate(BaseModel):
    name: Optional[str] = None
    entity_type: Optional[str] = None
    description: Optional[str] = None
    business_unit_id: Optional[int] = None
    audit_cycle_months: Optional[int] = None
    owner_id: Optional[int] = None
    status: Optional[str] = None
    industry: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_designation: Optional[str] = None
    metadata_json: Optional[dict] = None
    risk_factors: Optional[Dict[str, float]] = None


class AIDescriptionRequest(BaseModel):
    entity_name: str
    entity_type: str
    industry: Optional[str] = None


def score_to_rating(score: float) -> str:
    if score >= 80:
        return "critical"
    elif score >= 70:
        return "high"
    elif score >= 50:
        return "medium"
    return "low"


CATEGORY_TO_TYPE = {
    "technology": "technology",
    "compliance": "compliance",
    "operational": "process",
    "strategic": "process",
    "financial": "process",
    "third_party": "process",
    "project_change": "process",
}


def serialize_entity(e: AuditableEntity) -> dict:
    risk_ids = e.linked_risk_ids or []
    return {
        "id": e.id,
        "tenant_id": e.tenant_id,
        "name": e.name,
        "entity_type": e.entity_type,
        "description": e.description,
        "business_unit_id": e.business_unit_id,
        "business_unit_name": e.business_unit.name if e.business_unit else None,
        "risk_score": e.risk_score,
        "risk_rating": e.risk_rating,
        "audit_cycle_months": e.audit_cycle_months,
        "last_audited_date": e.last_audited_date.isoformat() if e.last_audited_date else None,
        "next_audit_due": e.next_audit_due.isoformat() if e.next_audit_due else None,
        "owner_id": e.owner_id,
        "owner_name": (e.owner.display_name or e.owner.username) if e.owner else None,
        "status": e.status,
        "industry": getattr(e, "industry", None),
        "contact_name": getattr(e, "contact_name", None),
        "contact_email": getattr(e, "contact_email", None),
        "contact_phone": getattr(e, "contact_phone", None),
        "contact_designation": getattr(e, "contact_designation", None),
        "linked_risk_ids": risk_ids,
        "linked_risk_count": len(risk_ids),
        "metadata_json": e.metadata_json,
        "auto_risk_score": getattr(e, "auto_risk_score", None),
        "risk_factors": getattr(e, "risk_factors", None) or {},
        "factor_contributions": getattr(e, "factor_contributions", None) or [],
        "scored_at": e.scored_at.isoformat() if getattr(e, "scored_at", None) else None,
        "score_override": bool(getattr(e, "score_override", False)),
        "override_score": getattr(e, "override_score", None),
        "override_rating": getattr(e, "override_rating", None),
        "override_justification": getattr(e, "override_justification", None),
        "override_by_id": getattr(e, "override_by_id", None),
        "override_at": e.override_at.isoformat() if getattr(e, "override_at", None) else None,
        "created_at": e.created_at.isoformat() if e.created_at else None,
        "updated_at": e.updated_at.isoformat() if e.updated_at else None,
    }


@router.get("")
def list_auditable_entities(
    entity_type: Optional[str] = None,
    risk_rating: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"entities": [], "total": 0}
    
    query = db.query(AuditableEntity).filter(AuditableEntity.tenant_id.in_(user_tenants))
    
    if entity_type:
        query = query.filter(AuditableEntity.entity_type == entity_type)
    if risk_rating:
        query = query.filter(AuditableEntity.risk_rating == risk_rating)
    if status:
        query = query.filter(AuditableEntity.status == status)
    
    entities = query.order_by(AuditableEntity.risk_score.desc()).all()
    return {"entities": [serialize_entity(e) for e in entities], "total": len(entities)}


@router.post("", status_code=status.HTTP_201_CREATED)
def create_auditable_entity(
    data: AuditableEntityCreate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=403, detail="No tenant access")
    
    next_due = datetime.utcnow() + timedelta(days=data.audit_cycle_months * 30)
    
    if data.industry and data.industry not in INDUSTRIES:
        raise HTTPException(status_code=422, detail=f"Invalid industry. Must be one of: {', '.join(INDUSTRIES)}")

    entity = AuditableEntity(
        tenant_id=tenant_id,
        name=data.name,
        entity_type=data.entity_type,
        description=data.description,
        business_unit_id=data.business_unit_id,
        risk_score=0,
        risk_rating="low",
        audit_cycle_months=data.audit_cycle_months,
        next_audit_due=next_due,
        owner_id=data.owner_id,
        industry=data.industry,
        contact_name=data.contact_name,
        contact_email=data.contact_email,
        contact_phone=data.contact_phone,
        contact_designation=data.contact_designation,
        metadata_json=data.metadata_json or {},
        risk_factors=_sanitize_factor_map(data.risk_factors),
    )
    db.add(entity)
    db.flush()
    # Score/rating are owned by the RBA engine, never set directly by clients.
    cfg = get_or_create_config(db, tenant_id)
    apply_score_to_entity(db, entity, normalize_weights(cfg.weights))
    db.commit()
    db.refresh(entity)
    return serialize_entity(entity)


def _sanitize_factor_map(raw: Optional[Dict[str, float]]) -> dict:
    """Keep only known manual factor keys, clamped to 0-100."""
    out: dict = {}
    if not raw:
        return out
    for key, value in raw.items():
        if key not in MANUAL_FACTOR_KEYS:
            continue
        try:
            out[key] = max(0.0, min(100.0, round(float(value), 1)))
        except (TypeError, ValueError):
            continue
    return out


def _get_openai_client():
    api_key = os.environ.get("AI_INTEGRATIONS_OPENAI_API_KEY")
    base_url = os.environ.get("AI_INTEGRATIONS_OPENAI_BASE_URL")
    if not api_key:
        raise HTTPException(status_code=503, detail="AI service not configured")
    from openai import OpenAI
    kwargs = {"api_key": api_key}
    if base_url:
        kwargs["base_url"] = base_url
    return OpenAI(**kwargs)


@router.post("/generate-description")
def generate_entity_description(
    data: AIDescriptionRequest,
    current_user: GRCUser = Depends(require_auth),
):
    if data.industry and data.industry not in INDUSTRIES:
        raise HTTPException(status_code=422, detail=f"Invalid industry. Must be one of: {', '.join(INDUSTRIES)}")

    client = _get_openai_client()
    industry_ctx = f" in the {data.industry} industry" if data.industry else ""
    prompt = (
        f"You are an enterprise GRC (Governance, Risk & Compliance) expert. "
        f"Generate a concise professional description for an auditable entity named "
        f"\"{data.entity_name}\" of type \"{data.entity_type}\"{industry_ctx}.\n\n"
        f"Cover the following in 3-5 sentences:\n"
        f"1. What this entity does and its role in the organization\n"
        f"2. Key business processes it supports\n"
        f"3. Regulatory relevance and compliance considerations\n"
        f"4. Inherent risk factors\n"
        f"5. Typical audit focus areas\n\n"
        f"Write in a professional tone suitable for an audit universe record. "
        f"Do not use bullet points or numbered lists — use flowing prose."
    )
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=500,
            temperature=0.7,
        )
        description = resp.choices[0].message.content.strip()
        return {"description": description}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI service error: {str(e)}")


# ---------------------------------------------------------------------------
# Bulk import of risk factors / financial materiality (spreadsheet / CSV)
# ---------------------------------------------------------------------------

IMPORT_IDENTITY_FIELDS = ["entity_id", "entity_name"]

_IMPORT_HEADER_ALIASES = {
    "entity_id": "entity_id", "id": "entity_id", "entity id": "entity_id",
    "entity_name": "entity_name", "name": "entity_name", "entity": "entity_name",
    "entity name": "entity_name", "auditable entity": "entity_name",
}


def _manual_factor_catalog() -> List[dict]:
    return [
        {"key": f["key"], "label": f["label"], "description": f["description"]}
        for f in FACTOR_DEFINITIONS if f["source"] == "manual"
    ]


def _default_import_mapping(headers: List[str]) -> Dict[str, Optional[str]]:
    """Auto-detect a column-header -> target mapping.

    Target is a manual factor key, an identity field (entity_id/entity_name),
    or None when the column is not recognized.
    """
    label_to_key = {
        f["label"].strip().lower(): f["key"]
        for f in FACTOR_DEFINITIONS if f["source"] == "manual"
    }
    manual_keys = set(MANUAL_FACTOR_KEYS)
    mapping: Dict[str, Optional[str]] = {}
    for h in headers:
        hl = str(h or "").strip().lower()
        if not hl:
            continue
        if hl in _IMPORT_HEADER_ALIASES:
            mapping[h] = _IMPORT_HEADER_ALIASES[hl]
        elif hl in manual_keys:
            mapping[h] = hl
        elif hl in label_to_key:
            mapping[h] = label_to_key[hl]
        else:
            mapping[h] = None
    return mapping


def _extract_universe_rows(file_name: str, file_bytes: bytes):
    """Return (headers, rows) where each row maps original-header -> raw value."""
    fn = (file_name or "").lower()
    headers: List[str] = []
    rows: List[dict] = []

    if fn.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="ignore")
        all_rows = list(csv.reader(text.splitlines()))
        if not all_rows:
            return [], []
        headers = [str(h or "").strip() for h in all_rows[0]]
        data_rows = all_rows[1:]
    elif fn.endswith((".xlsx", ".xls")):
        if openpyxl is None:
            raise HTTPException(status_code=500, detail="openpyxl is not installed")
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        headers = [str(h or "").strip() for h in all_rows[0]]
        data_rows = all_rows[1:]
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use .xlsx/.xls or .csv")

    for r in data_rows:
        if not r or not any(v is not None and str(v).strip() for v in r):
            continue
        item = {}
        for i, value in enumerate(r):
            if i < len(headers) and headers[i]:
                item[headers[i]] = value
        if item:
            rows.append(item)
    return headers, rows


def _resolve_mapping(headers: List[str], override) -> Dict[str, Optional[str]]:
    detected = _default_import_mapping(headers)
    if override:
        valid_targets = set(MANUAL_FACTOR_KEYS) | set(IMPORT_IDENTITY_FIELDS)
        for h, target in (override or {}).items():
            if h in detected:
                detected[h] = target if target in valid_targets else None
    return detected


def _resolve_import_rows(db: Session, user_tenants, rows: List[dict], mapping: Dict[str, Optional[str]]):
    manual_keys = set(MANUAL_FACTOR_KEYS)
    id_headers = [h for h, t in mapping.items() if t == "entity_id"]
    name_headers = [h for h, t in mapping.items() if t == "entity_name"]
    factor_headers = {h: t for h, t in mapping.items() if t in manual_keys}

    entities = db.query(AuditableEntity).filter(
        AuditableEntity.tenant_id.in_(user_tenants)
    ).all()
    by_id = {e.id: e for e in entities}
    by_name: Dict[str, AuditableEntity] = {}
    for e in entities:
        by_name.setdefault((e.name or "").strip().lower(), e)

    results = []
    for idx, row in enumerate(rows, start=2):
        errors: List[str] = []
        warnings: List[str] = []
        matched: Optional[AuditableEntity] = None
        identifier = None

        raw_id = next((row.get(h) for h in id_headers if row.get(h) not in (None, "")), None)
        if raw_id not in (None, ""):
            try:
                eid = int(float(str(raw_id).strip()))
                identifier = str(eid)
                matched = by_id.get(eid)
                if not matched:
                    errors.append(f"No entity with id {eid}")
            except (TypeError, ValueError):
                errors.append(f"Invalid entity_id '{raw_id}'")

        if matched is None and not errors:
            raw_name = next((row.get(h) for h in name_headers if row.get(h) not in (None, "")), None)
            if raw_name not in (None, ""):
                identifier = str(raw_name).strip()
                matched = by_name.get(identifier.lower())
                if not matched:
                    errors.append(f"No entity named '{identifier}'")
            else:
                errors.append("Row has no entity_id or entity_name")

        values: Dict[str, float] = {}
        for h, key in factor_headers.items():
            raw = row.get(h)
            if raw in (None, ""):
                continue
            try:
                v = float(str(raw).strip())
            except (TypeError, ValueError):
                errors.append(f"'{h}': not a number ('{raw}')")
                continue
            if v < 0 or v > 100:
                v = max(0.0, min(100.0, v))
                warnings.append(f"'{h}': clamped to {v:g}")
            values[key] = round(v, 1)

        if not values and not errors:
            warnings.append("No risk-factor values to apply")

        results.append({
            "row_number": idx,
            "identifier": identifier,
            "matched_entity_id": matched.id if matched else None,
            "matched_entity_name": matched.name if matched else None,
            "values": values,
            "errors": errors,
            "warnings": warnings,
            "valid": len(errors) == 0 and len(values) > 0,
        })
    return results


@router.get("/import/template")
def download_universe_import_template(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="No tenant access")
    if openpyxl is None:
        raise HTTPException(status_code=500, detail="openpyxl is not installed")

    manual = _manual_factor_catalog()
    headers = ["entity_id", "entity_name"] + [f["key"] for f in manual]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RiskFactors"
    ws.append(headers)

    sample = db.query(AuditableEntity).filter(
        AuditableEntity.tenant_id.in_(user_tenants)
    ).order_by(AuditableEntity.risk_score.desc().nullslast()).limit(5).all()
    if sample:
        for e in sample:
            rf = e.risk_factors or {}
            ws.append([e.id, e.name] + [rf.get(f["key"], "") for f in manual])
    else:
        ws.append(["", "Example Entity Name"] + [50 for _ in manual])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    file_name = f"risk_factors_template_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


def _parse_mapping_form(mapping: Optional[str]):
    if not mapping:
        return None
    try:
        return json.loads(mapping) or None
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON")


@router.post("/import/preview")
async def preview_universe_import(
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        raise HTTPException(status_code=403, detail="No tenant access")
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required")

    file_bytes = await file.read()
    headers, rows = _extract_universe_rows(file.filename, file_bytes)
    if not headers:
        raise HTTPException(status_code=400, detail="File has no header row")

    resolved_mapping = _resolve_mapping(headers, _parse_mapping_form(mapping))
    resolved_rows = _resolve_import_rows(db, user_tenants, rows, resolved_mapping)
    summary = {
        "total_rows": len(resolved_rows),
        "valid_rows": sum(1 for r in resolved_rows if r["valid"]),
        "matched_rows": sum(1 for r in resolved_rows if r["matched_entity_id"]),
        "error_rows": sum(1 for r in resolved_rows if r["errors"]),
    }
    return {
        "columns": headers,
        "mapping": resolved_mapping,
        "factor_options": _manual_factor_catalog(),
        "identity_options": IMPORT_IDENTITY_FIELDS,
        "rows": resolved_rows,
        "summary": summary,
    }


@router.post("/import/commit")
async def commit_universe_import(
    file: UploadFile = File(...),
    mapping: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth),
):
    user_tenants = get_user_tenants(current_user, db)
    tenant_id = get_user_primary_tenant(current_user, db)
    if not user_tenants or not tenant_id:
        raise HTTPException(status_code=403, detail="No tenant access")
    if not file.filename:
        raise HTTPException(status_code=400, detail="File is required")

    file_bytes = await file.read()
    headers, rows = _extract_universe_rows(file.filename, file_bytes)
    if not headers:
        raise HTTPException(status_code=400, detail="File has no header row")

    resolved_mapping = _resolve_mapping(headers, _parse_mapping_form(mapping))
    resolved_rows = _resolve_import_rows(db, user_tenants, rows, resolved_mapping)

    cfg = get_or_create_config(db, tenant_id)
    weights = normalize_weights(cfg.weights)
    updated = 0
    skipped = 0
    errors: List[str] = []

    for r in resolved_rows:
        if not r["valid"] or not r["matched_entity_id"]:
            skipped += 1
            if r["errors"]:
                errors.append(f"Row {r['row_number']}: {'; '.join(r['errors'])}")
            continue
        entity = db.query(AuditableEntity).filter(
            AuditableEntity.id == r["matched_entity_id"],
            AuditableEntity.tenant_id.in_(user_tenants),
        ).first()
        if not entity:
            skipped += 1
            continue
        factors = dict(entity.risk_factors or {})
        factors.update(r["values"])
        entity.risk_factors = factors
        rescore_entity(
            db, entity, weights, reason="import",
            user_id=current_user.id,
            alert_delta=cfg.alert_delta if cfg.alert_delta is not None else 10.0,
            alert_on_rating_change=cfg.alert_on_rating_change if cfg.alert_on_rating_change is not None else True,
        )
        updated += 1

    db.commit()
    return {
        "message": f"Updated {updated} entities ({skipped} skipped)",
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


@router.get("/coverage-gaps")
def get_coverage_gaps(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"gaps": [], "stats": {}}
    
    entities = db.query(AuditableEntity).filter(
        AuditableEntity.tenant_id.in_(user_tenants),
        AuditableEntity.status == "active"
    ).all()
    
    now = datetime.utcnow()
    never_audited = []
    overdue = []
    upcoming = []
    on_track = []
    
    for e in entities:
        if not e.last_audited_date:
            never_audited.append(serialize_entity(e))
        elif e.next_audit_due and e.next_audit_due < now:
            overdue.append(serialize_entity(e))
        elif e.next_audit_due and e.next_audit_due < now + timedelta(days=90):
            upcoming.append(serialize_entity(e))
        else:
            on_track.append(serialize_entity(e))
    
    total = len(entities)
    coverage_pct = ((total - len(never_audited)) / total * 100) if total > 0 else 0
    
    return {
        "gaps": {
            "never_audited": never_audited,
            "overdue": overdue,
            "upcoming_90_days": upcoming,
            "on_track": on_track,
        },
        "stats": {
            "total_entities": total,
            "never_audited_count": len(never_audited),
            "overdue_count": len(overdue),
            "upcoming_count": len(upcoming),
            "on_track_count": len(on_track),
            "coverage_percentage": round(coverage_pct, 1),
        }
    }


@router.get("/risk-enrichment")
def get_risk_enrichment(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    if not user_tenants:
        return {"risk_summary": [], "total_risks": 0}
    
    risks = db.query(
        Risk.category,
        func.count(Risk.id).label("count"),
        func.avg(Risk.residual_score).label("avg_score"),
        func.max(Risk.residual_score).label("max_score")
    ).filter(
        Risk.tenant_id.in_(user_tenants),
        Risk.status != "closed"
    ).group_by(Risk.category).all()
    
    total = sum(r.count for r in risks)
    
    entities = db.query(AuditableEntity).filter(
        AuditableEntity.tenant_id.in_(user_tenants)
    ).all()
    linked_count = sum(1 for e in entities if e.linked_risk_ids and len(e.linked_risk_ids) > 0)
    total_linked_risks = sum(len(e.linked_risk_ids or []) for e in entities)
    
    return {
        "risk_summary": [{
            "category": r.category,
            "count": r.count,
            "avg_score": round(float(r.avg_score or 0), 1),
            "max_score": float(r.max_score or 0),
        } for r in risks],
        "total_risks": total,
        "linked_entities": linked_count,
        "total_linked_risks": total_linked_risks,
        "total_entities": len(entities),
    }


@router.post("/sync-from-risks")
def sync_from_risk_register(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=403, detail="No tenant access")
    
    open_risks = db.query(Risk).filter(
        Risk.tenant_id == tenant_id,
        Risk.status != "closed"
    ).all()
    
    if not open_risks:
        return {"message": "No open risks found in Risk Register", "created": 0, "updated": 0}
    
    by_category = {}
    for risk in open_risks:
        cat = risk.category or "operational"
        if cat not in by_category:
            by_category[cat] = []
        by_category[cat].append(risk)
    
    created = 0
    updated = 0
    results = []
    result_refs: List[tuple] = []
    touched: List[AuditableEntity] = []
    
    for category, cat_risks in by_category.items():
        risk_ids = [r.id for r in cat_risks]
        max_score = max((r.residual_score or r.inherent_score or 0) for r in cat_risks)
        entity_type = CATEGORY_TO_TYPE.get(category, "process")
        entity_name = f"{category.replace('_', ' ').title()} Risk Area"
        
        existing = db.query(AuditableEntity).filter(
            AuditableEntity.tenant_id == tenant_id,
            AuditableEntity.name == entity_name,
        ).first()
        
        if existing:
            # Don't pre-seed risk_score/risk_rating: the rescore below derives the
            # true composite from factors and must compare against the real prior
            # baseline (pre-seeding would corrupt the delta and alert logic).
            existing.linked_risk_ids = risk_ids
            existing.updated_at = datetime.utcnow()
            updated += 1
            touched.append(existing)
            res = {"name": entity_name, "action": "updated", "risk_score": existing.risk_score, "risk_count": len(risk_ids)}
            results.append(res)
            result_refs.append((res, existing))
        else:
            # New entities start at score 0 so the first rescore establishes the
            # baseline without emitting a spurious first-run alert.
            entity = AuditableEntity(
                tenant_id=tenant_id,
                name=entity_name,
                entity_type=entity_type,
                description=f"Auto-generated from {len(cat_risks)} {category} risk(s) in the Risk Register",
                risk_score=0,
                risk_rating="low",
                audit_cycle_months=12 if max_score < 70 else 6,
                next_audit_due=datetime.utcnow() + timedelta(days=180 if max_score < 70 else 90),
                linked_risk_ids=risk_ids,
                status="active",
            )
            db.add(entity)
            created += 1
            touched.append(entity)
            res = {"name": entity_name, "action": "created", "risk_score": 0, "risk_count": len(risk_ids)}
            results.append(res)
            result_refs.append((res, entity))
    
    for risk in open_risks:
        score = risk.residual_score or risk.inherent_score or 0
        if score < 80:
            continue
        
        entity_name = risk.title
        existing = db.query(AuditableEntity).filter(
            AuditableEntity.tenant_id == tenant_id,
            AuditableEntity.name == entity_name,
        ).first()
        
        if existing:
            # See note above: never pre-seed the score before rescore.
            if not existing.linked_risk_ids or risk.id not in existing.linked_risk_ids:
                existing.linked_risk_ids = list(set((existing.linked_risk_ids or []) + [risk.id]))
            existing.updated_at = datetime.utcnow()
            updated += 1
            touched.append(existing)
            res = {"name": entity_name, "action": "updated", "risk_score": existing.risk_score, "risk_count": 1}
            results.append(res)
            result_refs.append((res, existing))
        else:
            cat = risk.category or "operational"
            entity = AuditableEntity(
                tenant_id=tenant_id,
                name=entity_name,
                entity_type=CATEGORY_TO_TYPE.get(cat, "process"),
                description=risk.description or f"Critical risk requiring dedicated audit coverage",
                risk_score=0,
                risk_rating="low",
                audit_cycle_months=6,
                next_audit_due=datetime.utcnow() + timedelta(days=90),
                linked_risk_ids=[risk.id],
                status="active",
            )
            db.add(entity)
            created += 1
            touched.append(entity)
            res = {"name": entity_name, "action": "created", "risk_score": 0, "risk_count": 1}
            results.append(res)
            result_refs.append((res, entity))

    # Derive the weighted composite from factors so synced entities carry true RBA
    # scores/ratings (not raw max-risk seeds). First rescore of a new entity sets
    # the baseline without alerting; updates compare against the real prior score.
    db.flush()
    cfg = get_or_create_config(db, tenant_id)
    weights = normalize_weights(cfg.weights)
    for ent in touched:
        rescore_entity(
            db, ent, weights, reason="sync_from_risks", user_id=current_user.id,
            alert_delta=cfg.alert_delta if cfg.alert_delta is not None else 10.0,
            alert_on_rating_change=cfg.alert_on_rating_change if cfg.alert_on_rating_change is not None else True,
        )

    # Report the rescored composite (not the transient seed) in the response.
    for res, ent in result_refs:
        res["risk_score"] = ent.risk_score

    db.commit()
    
    return {
        "message": f"Synced {created + updated} entities from Risk Register",
        "created": created,
        "updated": updated,
        "details": results,
    }


@router.post("/refresh-risk-scores")
def refresh_risk_scores(
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    tenant_id = get_user_primary_tenant(current_user, db)
    if not tenant_id:
        raise HTTPException(status_code=403, detail="No tenant access")
    
    entities = db.query(AuditableEntity).filter(
        AuditableEntity.tenant_id == tenant_id,
    ).all()

    cfg = get_or_create_config(db, tenant_id)
    weights = normalize_weights(cfg.weights)

    refreshed = 0
    alerts = 0
    for entity in entities:
        # Prune closed risks from the linkage so contributions stay accurate.
        risk_ids = entity.linked_risk_ids or []
        if risk_ids:
            linked_risks = db.query(Risk).filter(Risk.id.in_(risk_ids), Risk.tenant_id == tenant_id).all()
            active_ids = [r.id for r in linked_risks if r.status != "closed"]
            if active_ids != risk_ids:
                entity.linked_risk_ids = active_ids
        info = rescore_entity(
            db, entity, weights, reason="refresh_risk_scores", user_id=current_user.id,
            alert_delta=cfg.alert_delta if cfg.alert_delta is not None else 10.0,
            alert_on_rating_change=cfg.alert_on_rating_change if cfg.alert_on_rating_change is not None else True,
        )
        if info.get("alert"):
            alerts += 1
        refreshed += 1

    db.commit()

    return {
        "message": f"Recomputed composite risk scores for {refreshed} entities",
        "refreshed": refreshed,
        "total_entities": len(entities),
        "alerts_raised": alerts,
    }


@router.get("/{entity_id}")
def get_auditable_entity(
    entity_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    entity = db.query(AuditableEntity).filter(
        AuditableEntity.id == entity_id,
        AuditableEntity.tenant_id.in_(user_tenants)
    ).first()
    if not entity:
        raise HTTPException(status_code=404, detail="Entity not found")
    
    engagements = db.query(AuditEngagement).filter(
        AuditEngagement.auditable_entity_id == entity_id
    ).order_by(AuditEngagement.created_at.desc()).all()
    
    result = serialize_entity(entity)
    result["audit_history"] = [{
        "id": eng.id,
        "title": eng.title,
        "status": eng.status,
        "opinion": eng.opinion,
        "planned_start": eng.planned_start.isoformat() if eng.planned_start else None,
        "planned_end": eng.planned_end.isoformat() if eng.planned_end else None,
    } for eng in engagements]
    
    risk_ids = entity.linked_risk_ids or []
    if risk_ids:
        linked_risks = db.query(Risk).filter(Risk.id.in_(risk_ids), Risk.tenant_id.in_(user_tenants)).all()
        result["linked_risks"] = [{
            "id": r.id,
            "title": r.title,
            "category": r.category,
            "residual_score": r.residual_score,
            "inherent_score": r.inherent_score,
            "status": r.status,
            "risk_rating": score_to_rating(r.residual_score or r.inherent_score or 0),
        } for r in linked_risks]
    else:
        result["linked_risks"] = []
    
    return result


@router.put("/{entity_id}")
def update_auditable_entity(
    entity_id: int,
    data: AuditableEntityUpdate,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    entity = db.query(AuditableEntity).filter(
        AuditableEntity.id == entity_id,
        AuditableEntity.tenant_id.in_(user_tenants)
    ).first()
    if not entity:
        raise HTTPException(status_code=404, detail="Entity not found")

    if data.industry is not None and data.industry and data.industry not in INDUSTRIES:
        raise HTTPException(status_code=422, detail=f"Invalid industry. Must be one of: {', '.join(INDUSTRIES)}")

    update_fields = data.dict(exclude_unset=True)
    if "risk_factors" in update_fields:
        provided = update_fields.pop("risk_factors")
        merged = dict(entity.risk_factors or {})
        merged.update(_sanitize_factor_map(provided))
        entity.risk_factors = merged
    for field, value in update_fields.items():
        setattr(entity, field, value)
    
    if data.audit_cycle_months and entity.last_audited_date:
        entity.next_audit_due = entity.last_audited_date + timedelta(days=data.audit_cycle_months * 30)

    # Keep score/rating canonical: recompute via the RBA engine (respects any
    # active manual override). Clients cannot set risk_score/risk_rating directly.
    cfg = get_or_create_config(db, entity.tenant_id)
    rescore_entity(
        db, entity, normalize_weights(cfg.weights), reason="entity_updated",
        user_id=current_user.id,
        alert_delta=cfg.alert_delta if cfg.alert_delta is not None else 10.0,
        alert_on_rating_change=cfg.alert_on_rating_change if cfg.alert_on_rating_change is not None else True,
    )
    entity.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(entity)
    return serialize_entity(entity)


@router.delete("/{entity_id}")
def delete_auditable_entity(
    entity_id: int,
    db: Session = Depends(get_db),
    current_user: GRCUser = Depends(require_auth)
):
    user_tenants = get_user_tenants(current_user, db)
    entity = db.query(AuditableEntity).filter(
        AuditableEntity.id == entity_id,
        AuditableEntity.tenant_id.in_(user_tenants)
    ).first()
    if not entity:
        raise HTTPException(status_code=404, detail="Entity not found")
    
    db.delete(entity)
    db.commit()
    return {"message": "Entity deleted"}
